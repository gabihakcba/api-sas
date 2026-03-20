from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_DIR = (
    Path("/home/gabihak/Documents/Scout/Sistema de Administracion Scout")
    / "api-sas"
    / "public"
    / "documentos-xlsl"
)

GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HELP_TITLE_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")

DATE_HELP = "Fecha en texto con formato exacto dd/mm/yyyy. Ejemplo: 20/09/1999"
BOOLEAN_HELP = 'Acepta: si, no, true, false, 1, 0'
NUMERIC_HELP = "Solo numeros, sin espacios ni simbolos"


def text_required_formula(column: str, row: int) -> str:
    return f'LEN(TRIM({column}{row}))>0'


def row_has_data_formula(last_column: str, row: int) -> str:
    return f'COUNTA(A{row}:{last_column}{row})>0'


def boolean_valid_formula(column: str, row: int) -> str:
    cell = f"{column}{row}"
    normalized = f"LOWER(TRIM({cell}))"
    return (
        f'OR({cell}="",'
        f'{normalized}="si",'
        f'{normalized}="no",'
        f'{normalized}="true",'
        f'{normalized}="false",'
        f'{normalized}="1",'
        f'{normalized}="0")'
    )


def numeric_text_valid_formula(column: str, row: int) -> str:
    cell = f"{column}{row}"
    return (
        f'OR({cell}="",AND('
        f'ISTEXT({cell}),'
        f'LEN(TRIM({cell}))>0,'
        f'LEN({cell})=LEN(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({cell},"0",""),"1",""),"2",""),"3",""),"4",""))+'
        f'LEN(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({cell},"5",""),"6",""),"7",""),"8",""),"9",""))'
        f'))'
    )


def numeric_text_filled_valid_formula(column: str, row: int) -> str:
    cell = f"{column}{row}"
    return (
        f'AND({cell}<>"",'
        f'ISTEXT({cell}),'
        f'LEN(TRIM({cell}))>0,'
        f'LEN({cell})=LEN(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({cell},"0",""),"1",""),"2",""),"3",""),"4",""))+'
        f'LEN(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({cell},"5",""),"6",""),"7",""),"8",""),"9",""))'
        f')'
    )


def required_filled_formula(column: str, row: int) -> str:
    return f'LEN(TRIM({column}{row}))>0'


def date_text_valid_formula(column: str, row: int) -> str:
    cell = f"{column}{row}"
    return (
        f'OR({cell}="",AND('
        f'ISTEXT({cell}),'
        f'LEN({cell})=10,'
        f'MID({cell},3,1)="/",'
        f'MID({cell},6,1)="/",'
        f'ISNUMBER(--LEFT({cell},2)),'
        f'ISNUMBER(--MID({cell},4,2)),'
        f'ISNUMBER(--RIGHT({cell},4)),'
        f'TEXT(DATE(VALUE(RIGHT({cell},4)),VALUE(MID({cell},4,2)),VALUE(LEFT({cell},2))),"dd/mm/yyyy")={cell}'
        f'))'
    )


def date_text_filled_valid_formula(column: str, row: int) -> str:
    cell = f"{column}{row}"
    return (
        f'AND({cell}<>"",'
        f'ISTEXT({cell}),'
        f'LEN({cell})=10,'
        f'MID({cell},3,1)="/",'
        f'MID({cell},6,1)="/",'
        f'ISNUMBER(--LEFT({cell},2)),'
        f'ISNUMBER(--MID({cell},4,2)),'
        f'ISNUMBER(--RIGHT({cell},4)),'
        f'TEXT(DATE(VALUE(RIGHT({cell},4)),VALUE(MID({cell},4,2)),VALUE(LEFT({cell},2))),"dd/mm/yyyy")={cell}'
        f')'
    )


def add_validation_and_colors(
    ws,
    *,
    headers: list[str],
    date_columns: set[str],
    boolean_columns: set[str],
    numeric_columns: set[str],
    required_columns: set[str],
    max_rows: int = 500,
) -> None:
    header_to_column = {header: chr(65 + index) for index, header in enumerate(headers)}
    last_column = chr(64 + len(headers))

    for header, column in header_to_column.items():
        if header in date_columns:
            ws.column_dimensions[column].width = 18
            for row in range(2, max_rows + 1):
                ws[f"{column}{row}"].number_format = "@"

            validation = DataValidation(
                type="custom",
                formula1=date_text_valid_formula(column, 2),
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Fecha invalida",
                error="Usa el formato dd/mm/yyyy. Ejemplo: 20/09/1999",
                promptTitle="Fecha esperada",
                prompt="Escribi la fecha como texto en formato dd/mm/yyyy.",
            )
            validation.add(f"{column}2:{column}{max_rows}")
            ws.add_data_validation(validation)

            ws.conditional_formatting.add(
                f"{column}2:{column}{max_rows}",
                FormulaRule(
                    formula=[
                        f'AND({row_has_data_formula(last_column, 2)},{date_text_filled_valid_formula(column, 2)})'
                    ],
                    fill=GREEN_FILL,
                ),
            )
            ws.conditional_formatting.add(
                f"{column}2:{column}{max_rows}",
                FormulaRule(
                    formula=[
                        f'AND({row_has_data_formula(last_column, 2)},{column}2<>"",NOT({date_text_filled_valid_formula(column, 2)}))'
                    ],
                    fill=RED_FILL,
                ),
            )

        elif header in boolean_columns:
            ws.column_dimensions[column].width = 14
            validation = DataValidation(
                type="list",
                formula1='"si,no,true,false,1,0"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Valor invalido",
                error='Usa uno de estos valores: si, no, true, false, 1, 0',
                promptTitle="Valores permitidos",
                prompt='Acepta: si, no, true, false, 1, 0',
            )
            validation.add(f"{column}2:{column}{max_rows}")
            ws.add_data_validation(validation)

            ws.conditional_formatting.add(
                f"{column}2:{column}{max_rows}",
                FormulaRule(
                    formula=[
                        f'AND({row_has_data_formula(last_column, 2)},{column}2<>"",{boolean_valid_formula(column, 2)})'
                    ],
                    fill=GREEN_FILL,
                ),
            )
            ws.conditional_formatting.add(
                f"{column}2:{column}{max_rows}",
                FormulaRule(
                    formula=[
                        f'AND({row_has_data_formula(last_column, 2)},{column}2<>"",NOT({boolean_valid_formula(column, 2)}))'
                    ],
                    fill=RED_FILL,
                ),
            )

        elif header in numeric_columns:
            ws.column_dimensions[column].width = 18
            for row in range(2, max_rows + 1):
                ws[f"{column}{row}"].number_format = "@"

            validation = DataValidation(
                type="custom",
                formula1=numeric_text_valid_formula(column, 2),
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Valor invalido",
                error="Usa solo numeros, sin espacios ni simbolos.",
                promptTitle="Solo numeros",
                prompt="Escribi solo numeros, por ejemplo 40111222 o 1122334455.",
            )
            validation.add(f"{column}2:{column}{max_rows}")
            ws.add_data_validation(validation)

            ws.conditional_formatting.add(
                f"{column}2:{column}{max_rows}",
                FormulaRule(
                    formula=[
                        f'AND({row_has_data_formula(last_column, 2)},{numeric_text_filled_valid_formula(column, 2)})'
                    ],
                    fill=GREEN_FILL,
                ),
            )
            ws.conditional_formatting.add(
                f"{column}2:{column}{max_rows}",
                FormulaRule(
                    formula=[
                        f'AND({row_has_data_formula(last_column, 2)},{column}2<>"",NOT({numeric_text_filled_valid_formula(column, 2)}))'
                    ],
                    fill=RED_FILL,
                ),
            )

        elif header in required_columns:
            ws.column_dimensions[column].width = 22
            ws.conditional_formatting.add(
                f"{column}2:{column}{max_rows}",
                FormulaRule(
                    formula=[
                        f'AND({row_has_data_formula(last_column, 2)},{required_filled_formula(column, 2)})'
                    ],
                    fill=GREEN_FILL,
                ),
            )
            ws.conditional_formatting.add(
                f"{column}2:{column}{max_rows}",
                FormulaRule(
                    formula=[
                        f'AND({row_has_data_formula(last_column, 2)},LEN(TRIM({column}2))=0)'
                    ],
                    fill=RED_FILL,
                ),
            )


def build_help_sheet(
    wb: Workbook,
    *,
    title: str,
    intro_lines: list[str],
    headers: list[str],
    example_row: list[str],
    field_help: dict[str, str],
) -> None:
    ws = wb.create_sheet("Ayuda")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = HELP_TITLE_FILL

    help_row = 2
    for line in intro_lines:
        ws[f"A{help_row}"] = line
        help_row += 1

    help_row += 1
    ws[f"A{help_row}"] = "Campos incluidos"
    ws[f"A{help_row}"].font = Font(bold=True)
    help_row += 1

    ws[f"A{help_row}"] = "Campo"
    ws[f"B{help_row}"] = "Regla"
    ws[f"A{help_row}"].font = Font(bold=True)
    ws[f"B{help_row}"].font = Font(bold=True)
    help_row += 1

    for header in headers:
        ws[f"A{help_row}"] = header
        ws[f"B{help_row}"] = field_help.get(header, "Texto opcional")
        help_row += 1

    help_row += 1
    ws[f"A{help_row}"] = "Fila de ejemplo"
    ws[f"A{help_row}"].font = Font(bold=True)
    help_row += 1

    for index, header in enumerate(headers, start=1):
        ws.cell(row=help_row, column=index, value=header).font = Font(bold=True)
        ws.cell(row=help_row + 1, column=index, value=example_row[index - 1])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 72


def build_template(
    *,
    filename: str,
    title: str,
    headers: list[str],
    example_row: list[str],
    date_columns: set[str],
    boolean_columns: set[str],
    numeric_columns: set[str],
    required_columns: set[str],
    intro_lines: list[str],
    field_help: dict[str, str],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Carga"

    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        ws.column_dimensions[chr(64 + index)].width = 22

    ws.freeze_panes = "A2"
    add_validation_and_colors(
        ws,
        headers=headers,
        date_columns=date_columns,
        boolean_columns=boolean_columns,
        numeric_columns=numeric_columns,
        required_columns=required_columns,
    )
    for index, value in enumerate(example_row, start=1):
        ws.cell(row=2, column=index, value=value)

    build_help_sheet(
        wb,
        title=title,
        intro_lines=intro_lines,
        headers=headers,
        example_row=example_row,
        field_help=field_help,
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_DIR / filename)


COMMON_PERSONAL_HEADERS = [
    "nombre",
    "apellidos",
    "dni",
    "fechaNacimiento",
    "direccion",
    "telefonoEmergencia",
    "email",
    "telefono",
    "totem",
    "cualidad",
]


build_template(
    filename="plantilla-import-protagonistas.xlsx",
    title="Plantilla Importacion Protagonistas",
    headers=COMMON_PERSONAL_HEADERS
    + ["fechaIngresoRama", "esBecado", "activo"],
    example_row=[
        "Juan",
        "Perez",
        "40111222",
        "15/03/2010",
        "Calle 123",
        "1122334455",
        "juan@mail.com",
        "1199988877",
        "Fenix",
        "Alegre",
        "01/03/2025",
        "no",
        "si",
    ],
    date_columns={"fechaNacimiento", "fechaIngresoRama"},
    boolean_columns={"esBecado", "activo"},
    numeric_columns={"dni", "telefono", "telefonoEmergencia"},
    required_columns={"nombre", "apellidos", "dni", "direccion"},
    intro_lines=[
        "La hoja Carga empieza vacia para que puedas pegar o editar filas desde la linea 2.",
        "No incluyas rama, user ni password. DNI se usa como usuario y contrasena.",
        "La rama se resuelve por quien importa o por la seleccion del dialogo.",
    ],
    field_help={
        "nombre": "Obligatorio",
        "apellidos": "Obligatorio",
        "dni": f"Obligatorio. Tambien se usa como user y password. {NUMERIC_HELP}",
        "fechaNacimiento": DATE_HELP,
        "direccion": "Obligatorio",
        "telefonoEmergencia": NUMERIC_HELP,
        "telefono": NUMERIC_HELP,
        "fechaIngresoRama": DATE_HELP,
        "esBecado": BOOLEAN_HELP,
        "activo": BOOLEAN_HELP,
    },
)

build_template(
    filename="plantilla-import-adultos.xlsx",
    title="Plantilla Importacion Adultos",
    headers=COMMON_PERSONAL_HEADERS + ["fechaInicioEquipo", "esBecado", "activo"],
    example_row=[
        "Maria",
        "Gomez",
        "30111222",
        "20/09/1990",
        "Avenida 456",
        "1166677788",
        "maria@mail.com",
        "1177712345",
        "Ceibo",
        "Constancia",
        "01/03/2025",
        "no",
        "si",
    ],
    date_columns={"fechaNacimiento", "fechaInicioEquipo"},
    boolean_columns={"esBecado", "activo"},
    numeric_columns={"dni", "telefono", "telefonoEmergencia"},
    required_columns={"nombre", "apellidos", "dni", "direccion"},
    intro_lines=[
        "La hoja Carga empieza vacia para que puedas pegar o editar filas desde la linea 2.",
        "No incluyas rama, area, posicion, role ni scope.",
        "DNI se usa como usuario y contrasena.",
        "Si se importa por planilla, se crea como AYUDANTE_RAMA en la rama resuelta o elegida.",
    ],
    field_help={
        "nombre": "Obligatorio",
        "apellidos": "Obligatorio",
        "dni": f"Obligatorio. Tambien se usa como user y password. {NUMERIC_HELP}",
        "fechaNacimiento": DATE_HELP,
        "direccion": "Obligatorio",
        "telefonoEmergencia": NUMERIC_HELP,
        "telefono": NUMERIC_HELP,
        "fechaInicioEquipo": DATE_HELP,
        "esBecado": BOOLEAN_HELP,
        "activo": BOOLEAN_HELP,
    },
)

build_template(
    filename="plantilla-import-responsables.xlsx",
    title="Plantilla Importacion Responsables",
    headers=COMMON_PERSONAL_HEADERS,
    example_row=[
        "Laura",
        "Fernandez",
        "28123456",
        "05/11/1987",
        "Pasaje 789",
        "1144455566",
        "laura@mail.com",
        "1161122233",
        "Colibri",
        "Escucha",
    ],
    date_columns={"fechaNacimiento"},
    boolean_columns=set(),
    numeric_columns={"dni", "telefono", "telefonoEmergencia"},
    required_columns={"nombre", "apellidos", "dni", "direccion"},
    intro_lines=[
        "La hoja Carga empieza vacia para que puedas pegar o editar filas desde la linea 2.",
        "No incluyas responsabilidades, user ni password.",
        "DNI se usa como usuario y contrasena.",
        "Las responsabilidades se asignan manualmente despues de importar.",
    ],
    field_help={
        "nombre": "Obligatorio",
        "apellidos": "Obligatorio",
        "dni": f"Obligatorio. Tambien se usa como user y password. {NUMERIC_HELP}",
        "fechaNacimiento": DATE_HELP,
        "direccion": "Obligatorio",
        "telefonoEmergencia": NUMERIC_HELP,
        "telefono": NUMERIC_HELP,
    },
)
